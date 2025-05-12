import openpyxl
import os
from openpyxl.utils import column_index_from_string

def append_to_pre(file_path, a_value, b_value, bs_value):
    if not file_path:

     os.makedirs(os.path.dirname(file_path), exist_ok=True)

    file_exists = os.path.exists(file_path)
    if file_exists:
        wb = openpyxl.load_workbook(file_path)
    else:
        wb = openpyxl.Workbook()
    ws = wb.active

    next_row = ws.max_row + 1 if file_exists else 1

    ws.cell(row=next_row, column=1, value=a_value)
    ws.cell(row=next_row, column=2, value=b_value)


    bs_col = column_index_from_string("BS")

    ws.cell(row=next_row, column=bs_col, value=bs_value)

    wb.save(file_path)

try:
    a = 1
    b = 74
    bs = 3
except Exception:
    raise ValueError("Variables A, B o BS no encontradas en el JSON")

pre_path = os.path.expanduser('~/importaciones/PRE.xlsx')


try:
    append_to_pre(pre_path, a, b, bs)
except Exception as e:
 print(f"Error al actualizar PRE.xlsx: {e}")